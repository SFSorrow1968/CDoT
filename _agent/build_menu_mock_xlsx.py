#!/usr/bin/env python3
"""
Generates MENU_MOCK.xlsx from CDoTModOptions.cs
Parses ModOption attributes to build a spreadsheet of all menu options.
"""

import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
MOD_OPTIONS_PATH = os.path.join(PROJECT_ROOT, "Configuration", "CDoTModOptions.cs")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "MENU_MOCK.xlsx")


def parse_constants(content):
    """Parse constant definitions from the file."""
    constants = {}
    # Match const int patterns like: private const int CategoryOrderPreset = 10;
    pattern = r'(?:private|public)?\s*const\s+int\s+(\w+)\s*=\s*(\d+)\s*;'
    for match in re.finditer(pattern, content):
        constants[match.group(1)] = int(match.group(2))
    return constants


def parse_mod_options(filepath):
    """Parse CDoTModOptions.cs and extract ModOption attributes."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Parse constants first
    constants = parse_constants(content)

    options = []

    # Match ModOption attributes followed by field declarations
    pattern = r'\[ModOption\(([^]]+)\)\]\s*public\s+static\s+(\w+)\s+(\w+)\s*=\s*([^;]+);'

    for match in re.finditer(pattern, content):
        attr_content = match.group(1)
        field_type = match.group(2)
        field_name = match.group(3)
        default_value = match.group(4).strip()

        option = {
            'field_name': field_name,
            'field_type': field_type,
            'default_value': default_value,
            'name': '',
            'category': 'Main',
            'categoryOrder': 0,
            'order': 0,
            'tooltip': '',
            'valueSourceName': '',
            'defaultValueIndex': 0,
        }

        # Parse attribute parameters
        name_match = re.search(r'name\s*=\s*(\w+)', attr_content)
        if name_match:
            option['name'] = name_match.group(1)

        category_match = re.search(r'category\s*=\s*(\w+)', attr_content)
        if category_match:
            option['category'] = category_match.group(1)

        category_order_match = re.search(r'categoryOrder\s*=\s*(\w+)', attr_content)
        if category_order_match:
            order_val = category_order_match.group(1)
            # Try to resolve constant, or use 0
            option['categoryOrder'] = constants.get(order_val, 0)

        order_match = re.search(r'order\s*=\s*(\d+)', attr_content)
        if order_match:
            option['order'] = int(order_match.group(1))

        tooltip_match = re.search(r'tooltip\s*=\s*"([^"]*)"', attr_content)
        if tooltip_match:
            option['tooltip'] = tooltip_match.group(1)

        source_match = re.search(r'valueSourceName\s*=\s*"(\w+)"', attr_content)
        if source_match:
            option['valueSourceName'] = source_match.group(1)

        default_idx_match = re.search(r'defaultValueIndex\s*=\s*(\d+)', attr_content)
        if default_idx_match:
            option['defaultValueIndex'] = int(default_idx_match.group(1))

        options.append(option)

    return options


def create_xlsx(options, output_path):
    """Create MENU_MOCK.xlsx from parsed options."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menu Options"

    # Headers
    headers = ['Category', 'Order', 'Name', 'Type', 'Default', 'Tooltip', 'Value Source']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Sort options by category order, then by order
    options.sort(key=lambda x: (x.get('categoryOrder', 0), x.get('order', 0)))

    # Add data rows
    for row, opt in enumerate(options, 2):
        ws.cell(row=row, column=1, value=opt['category'])
        ws.cell(row=row, column=2, value=opt['order'])
        ws.cell(row=row, column=3, value=opt['name'])
        ws.cell(row=row, column=4, value=opt['field_type'])
        ws.cell(row=row, column=5, value=opt['default_value'])
        ws.cell(row=row, column=6, value=opt['tooltip'])
        ws.cell(row=row, column=7, value=opt['valueSourceName'])

    # Adjust column widths
    column_widths = [20, 8, 25, 10, 15, 50, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(output_path)
    print(f"Generated: {output_path}")
    print(f"Total options: {len(options)}")


def main():
    if not os.path.exists(MOD_OPTIONS_PATH):
        print(f"Error: {MOD_OPTIONS_PATH} not found")
        sys.exit(1)

    options = parse_mod_options(MOD_OPTIONS_PATH)
    create_xlsx(options, OUTPUT_PATH)


if __name__ == "__main__":
    main()
