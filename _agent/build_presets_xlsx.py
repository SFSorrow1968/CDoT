#!/usr/bin/env python3
"""
Generates PRESETS.xlsx showing all preset values for each zone.
This shows exactly what values are applied when each preset is selected.
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "PRESETS.xlsx")

# Zone order
ZONES = ['Throat', 'Head', 'Neck', 'Torso', 'Arm', 'Leg', 'Dismemberment']

# ========== DAMAGE PRESET VALUES ==========
# 5 presets: Minimal (0), Low (1), Default (2), High (3), Extreme (4)
DAMAGE_PRESETS = ['Minimal', 'Low', 'Default', 'High', 'Extreme']
DAMAGE_VALUES = {
    'Throat':        [0.5,  1.25, 2.5,  5.0,  10.0],
    'Head':          [0.25, 0.75, 1.5,  3.0,  6.0],
    'Neck':          [0.5,  1.0,  2.0,  4.0,  8.0],
    'Torso':         [0.25, 0.5,  1.0,  2.0,  4.0],
    'Arm':           [0.25, 0.25, 0.5,  1.0,  2.0],
    'Leg':           [0.25, 0.5,  0.75, 1.5,  3.0],
    'Dismemberment': [1.0,  2.0,  3.0,  6.0,  12.0],
}

# ========== DURATION PRESET VALUES ==========
# 5 presets: VeryShort (0), Short (1), Default (2), Long (3), Extended (4)
DURATION_PRESETS = ['Very Short', 'Short', 'Default', 'Long', 'Extended']
DURATION_VALUES = {
    'Throat':        [2.0,  4.0,  6.0,  10.0, 15.0],
    'Head':          [1.5,  3.0,  5.0,  8.0,  12.0],
    'Neck':          [2.0,  3.5,  5.5,  9.0,  14.0],
    'Torso':         [1.5,  2.5,  4.0,  7.0,  10.0],
    'Arm':           [1.0,  2.0,  3.0,  5.0,  8.0],
    'Leg':           [1.0,  2.5,  3.5,  6.0,  9.0],
    'Dismemberment': [3.0,  5.0,  8.0,  12.0, 20.0],
}

# ========== FREQUENCY PRESET VALUES (per-zone) ==========
# 5 presets: VerySlow (0), Slow (1), Default (2), Fast (3), Rapid (4)
# Values are tick intervals in seconds (lower = faster ticks)
FREQUENCY_PRESETS = ['Very Slow', 'Slow', 'Default', 'Fast', 'Rapid']
FREQUENCY_VALUES = {
    'Throat':        [2.0,  1.0,  0.5,  0.3,  0.1],
    'Head':          [2.5,  1.2,  0.6,  0.3,  0.1],
    'Neck':          [2.0,  1.0,  0.5,  0.25, 0.1],
    'Torso':         [3.0,  1.5,  0.8,  0.4,  0.2],
    'Arm':           [3.5,  1.8,  1.0,  0.5,  0.2],
    'Leg':           [3.0,  1.5,  0.8,  0.4,  0.2],
    'Dismemberment': [1.5,  0.8,  0.4,  0.2,  0.1],
}

# ========== CHANCE PRESET VALUES ==========
# 5 presets: Off (0), Rare (1), Default (2), Frequent (3), Always (4)
CHANCE_PRESETS = ['Off', 'Rare', 'Default', 'Frequent', 'Always']
CHANCE_VALUES = {
    'Throat':        [0,  30, 60, 85,  100],
    'Head':          [0,  20, 40, 65,  100],
    'Neck':          [0,  25, 55, 80,  100],
    'Torso':         [0,  15, 35, 55,  100],
    'Arm':           [0,  10, 25, 45,  100],
    'Leg':           [0,  15, 30, 50,  100],
    'Dismemberment': [0,  40, 80, 95,  100],
}

# Styling
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DEFAULT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for Default
ZONE_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_sheet(wb, sheet_name, presets, values_dict, unit='', format_func=None):
    """Create a sheet for a preset category with per-zone values."""
    ws = wb.create_sheet(title=sheet_name)
    
    # Header row
    ws.cell(row=1, column=1, value="Zone").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).border = THIN_BORDER
    
    for col, preset in enumerate(presets, start=2):
        cell = ws.cell(row=1, column=col, value=preset)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Per-zone values
    for row, zone in enumerate(ZONES, start=2):
        ws.cell(row=row, column=1, value=zone).font = ZONE_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        
        values = values_dict[zone]
        for col, val in enumerate(values, start=2):
            if format_func:
                display = format_func(val)
            else:
                display = f"{val}{unit}"
            cell = ws.cell(row=row, column=col, value=display)
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            if col == 4:  # Default is index 2, column 4
                cell.fill = DEFAULT_FILL
    
    # Auto-width columns
    for col in range(1, len(presets) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12


def create_xlsx(output_path):
    """Create PRESETS.xlsx with all preset configurations."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    del wb['Sheet']
    
    # Create sheets for each preset category
    create_sheet(wb, "Damage", DAMAGE_PRESETS, DAMAGE_VALUES, unit='')
    create_sheet(wb, "Duration", DURATION_PRESETS, DURATION_VALUES, unit='s')
    create_sheet(wb, "Frequency", FREQUENCY_PRESETS, FREQUENCY_VALUES, unit='s')
    create_sheet(wb, "Chance", CHANCE_PRESETS, CHANCE_VALUES, unit='%')
    
    # Create summary sheet
    ws = wb.create_sheet(title="Summary", index=0)
    ws.cell(row=1, column=1, value="CDoT Preset System").font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Each preset category has 5 levels with per-zone values:")
    ws.cell(row=4, column=1, value="  • Default (middle) is always index 2")
    ws.cell(row=5, column=1, value="  • 2 presets to the left (lower/slower values)")
    ws.cell(row=6, column=1, value="  • 2 presets to the right (higher/faster values)")
    ws.cell(row=8, column=1, value="Damage Presets:").font = Font(bold=True)
    ws.cell(row=9, column=1, value="  Minimal → Low → Default → High → Extreme")
    ws.cell(row=11, column=1, value="Duration Presets:").font = Font(bold=True)
    ws.cell(row=12, column=1, value="  Very Short → Short → Default → Long → Extended")
    ws.cell(row=14, column=1, value="Frequency Presets (per-zone tick intervals):").font = Font(bold=True)
    ws.cell(row=15, column=1, value="  Very Slow → Slow → Default → Fast → Rapid")
    ws.cell(row=16, column=1, value="  (Each zone has unique tick intervals, e.g., Throat: 2.0s → 0.1s)")
    ws.cell(row=17, column=1, value="  Slider range: 0.1s to 5.0s in 0.1s increments")
    ws.cell(row=19, column=1, value="Chance Presets:").font = Font(bold=True)
    ws.cell(row=20, column=1, value="  Off → Rare → Default → Frequent → Always")
    ws.cell(row=22, column=1, value="Damage Type Multipliers:").font = Font(bold=True)
    ws.cell(row=23, column=1, value="  Pierce: 1.2x (default)")
    ws.cell(row=24, column=1, value="  Slash: 0.8x (default)")
    ws.cell(row=25, column=1, value="  Fire: 0.3x (default)")
    ws.cell(row=26, column=1, value="  Lightning: 1.5x (default)")
    ws.cell(row=27, column=1, value="  Set to 0.0x to disable DOT from that damage type")
    ws.cell(row=28, column=1, value="  Note: Blunt damage does not cause bleeding")
    
    ws.column_dimensions['A'].width = 60
    
    wb.save(output_path)
    print(f"Generated: {output_path}")


if __name__ == "__main__":
    create_xlsx(OUTPUT_PATH)
